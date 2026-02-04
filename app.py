from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
import os

app = Flask(__name__)
CORS(app)  # CORS ni yoqish - frontend dan so'rov yuborish uchun

# Admin paroli
ADMIN_PASSWORD = 'admin123'

# Database yaratish va jadval tuzish
def init_db():
    conn = sqlite3.connect('hisobot.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            foydalanuvchi TEXT NOT NULL,
            sana TEXT NOT NULL,
            ertalab TEXT,
            kechqurun TEXT,
            yuborilgan_vaqti TEXT NOT NULL,
            farq TEXT
        )
    ''')
    
    conn.commit()
    conn.close()

# Serverni ishga tushirganda database yaratish
init_db()

# Ma'lumot qo'shish endpoint
@app.route('/api/add-entry', methods=['POST'])
def add_entry():
    try:
        data = request.json
        
        foydalanuvchi = data.get('foydalanuvchi')
        sana = data.get('sana')
        muddat = data.get('muddat')  # ertalab yoki kechqurun
        raqam = data.get('raqam')
        
        # Hozirgi vaqt
        yuborilgan_vaqti = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        
        # Ertalab yoki Kechqurun ga raqam qo'yish
        ertalab = raqam if muddat == 'ertalab' else ''
        kechqurun = raqam if muddat == 'kechqurun' else ''
        
        conn = sqlite3.connect('hisobot.db')
        cursor = conn.cursor()
        
        # Shu foydalanuvchi va sana uchun mavjud yozuv bormi tekshirish
        cursor.execute('''
            SELECT id, ertalab, kechqurun FROM entries 
            WHERE foydalanuvchi = ? AND sana = ?
            ORDER BY id DESC LIMIT 1
        ''', (foydalanuvchi, sana))
        
        existing = cursor.fetchone()
        
        if existing:
            # Agar mavjud yozuv bo'lsa, uni yangilash
            existing_id, existing_ertalab, existing_kechqurun = existing
            
            # Yangi qiymatlarni birlashtirish
            final_ertalab = ertalab if ertalab else existing_ertalab
            final_kechqurun = kechqurun if kechqurun else existing_kechqurun
            
            # Farqni hisoblash
            farq = ''
            if final_ertalab and final_kechqurun:
                try:
                    farq = str(int(final_kechqurun) - int(final_ertalab))
                except:
                    farq = ''
            
            # Mavjud yozuvni yangilash
            cursor.execute('''
                UPDATE entries 
                SET ertalab = ?, kechqurun = ?, yuborilgan_vaqti = ?, farq = ?
                WHERE id = ?
            ''', (final_ertalab, final_kechqurun, yuborilgan_vaqti, farq, existing_id))
        else:
            # Yangi yozuv qo'shish
            cursor.execute('''
                INSERT INTO entries (foydalanuvchi, sana, ertalab, kechqurun, yuborilgan_vaqti, farq)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (foydalanuvchi, sana, ertalab, kechqurun, yuborilgan_vaqti, ''))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Ma\'lumot saqlandi'}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Admin login tekshirish
@app.route('/api/admin-login', methods=['POST'])
def admin_login():
    try:
        data = request.json
        password = data.get('password')
        
        if password == ADMIN_PASSWORD:
            return jsonify({'success': True}), 200
        else:
            return jsonify({'success': False, 'message': 'Noto\'g\'ri parol'}), 401
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Ma'lumotlarni olish (filtrlash bilan)
@app.route('/api/get-entries', methods=['POST'])
def get_entries():
    try:
        data = request.json
        
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        users = data.get('users', [])  # List
        
        conn = sqlite3.connect('hisobot.db')
        cursor = conn.cursor()
        
        # SQL query tuzish
        query = 'SELECT * FROM entries WHERE 1=1'
        params = []
        
        if start_date:
            query += ' AND sana >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND sana <= ?'
            params.append(end_date)
        
        if users and len(users) > 0:
            placeholders = ','.join(['?'] * len(users))
            query += f' AND foydalanuvchi IN ({placeholders})'
            params.extend(users)
        
        query += ' ORDER BY yuborilgan_vaqti DESC'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # Ma'lumotlarni formatga keltirish
        entries = []
        for row in rows:
            entries.append({
                'id': row[0],
                'foydalanuvchi': row[1],
                'sana': row[2],
                'ertalab': row[3],
                'kechqurun': row[4],
                'yuborilgan_vaqti': row[5],
                'farq': row[6]
            })
        
        conn.close()
        
        return jsonify({'success': True, 'entries': entries}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Excel yuklab olish
@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    try:
        data = request.json
        
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        users = data.get('users', [])
        
        # Ma'lumotlarni olish
        conn = sqlite3.connect('hisobot.db')
        cursor = conn.cursor()
        
        query = 'SELECT * FROM entries WHERE 1=1'
        params = []
        
        if start_date:
            query += ' AND sana >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND sana <= ?'
            params.append(end_date)
        
        if users and len(users) > 0:
            placeholders = ','.join(['?'] * len(users))
            query += f' AND foydalanuvchi IN ({placeholders})'
            params.extend(users)
        
        query += ' ORDER BY yuborilgan_vaqti DESC'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        # Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hisobot"
        
        # Header
        headers = ['Foydalanuvchi', 'Sana', 'Ertalab', 'Kechqurun', 'Yuborilgan vaqti', 'Farq']
        ws.append(headers)
        
        # Header stilini o'rnatish
        header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ma'lumotlarni qo'shish
        for row in rows:
            ws.append([row[1], row[2], row[3], row[4], row[5], row[6]])
        
        # Ustunlar kengligini o'rnatish
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        
        # Excel ni xotirada saqlash
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Fayl nomini yaratish
        filename = f'hisobot_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Server ishga tushirish
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)